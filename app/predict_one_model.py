import pandas as pd
import numpy as np
import pickle
import json
from app.my_libs.calc_features import *
from sklearn.preprocessing import StandardScaler
from datetime import datetime
import os
import warnings
warnings.filterwarnings("ignore")
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from app.my_libs.calc_features import *

ls_columns_result = [
    'Врем. сопротивление',
    'прогноз Врем. сопротивление',
    'MAE Врем. сопротивление',
    #     'MSE Врем. сопротивление',
    'Предел текучести',
    'прогноз Предел текучести',
    'MAE Предел текучести',
    #     'MSE Предел текучести',
]

ls_columns_output = [
    # 'Дата термообработки',
    '№ партии',
    'марка стали',
    'диаметр',
    'толщина стенки',
    'Гр. прочн.',
    '1 зона по ВТР закалка',
    '2 зона по ВТР закалка',
    '3 зона по ВТР закалка',
    'шаг балок закалочная печь, сек',
    'Скорость прохождения трубы через спрейер, м/с',
    't˚ C трубы после спреера',
    '1 зона ВТР и уставка отпуск',
    '2 зона ВТР и уставка отпуск',
    '3 зона ВТР и уставка отпуск',
    '4 зона ВТР и уставка отпуск',
    '5 зона ВТР и уставка отпуск',
    'шаг балок отпускная печь, сек',
    'C',
    'Mn',
    'Si',
    'P',
    'S',
    'Cr',
    'Ni',
    'Cu',
    'Al',
    'V',
    'Ti',
    'Nb',
    'Mo',
    'N',
    'B',
    'C-coef',
    'Параметр закалка',
    'Параметр отпуск',
    'Параметр отпуск новый',
    'Параметр отпуск новый 2',
    'Параметр отпуск новый V',
    'Величина зерна',
    'Тип предела текучести (1186)',
    'ICD'
]

ls_columns_required = [
    'марка стали',
    'диаметр',
    'толщина стенки',
    'Гр. прочн.',
    '1 зона по ВТР закалка',
    '2 зона по ВТР закалка',
    '3 зона по ВТР закалка',
    'шаг балок закалочная печь, сек',
    'Скорость прохождения трубы через спрейер, м/с',
    't˚ C трубы после спреера',
    '1 зона ВТР и уставка отпуск',
    '2 зона ВТР и уставка отпуск',
    '3 зона ВТР и уставка отпуск',
    '4 зона ВТР и уставка отпуск',
    '5 зона ВТР и уставка отпуск',
    'шаг балок отпускная печь, сек',
    'Тип предела текучести (1186)'
]

ls_opt_need = ['Гр. прочн.', 'марка стали', 'толщина стенки', 'диаметр']

#ADD START
col_out = ['Примечание', 'Врем. сопротивление', 'прогноз Врем. сопротивление',
       'MAE Врем. сопротивление', 'Предел текучести', 'прогноз Предел текучести', 'MAE Предел текучести', '№ партии', 'марка стали', 'диаметр',
       'толщина стенки', 'Гр. прочн.', '1 зона по ВТР закалка',
       '2 зона по ВТР закалка', '3 зона по ВТР закалка',
       'шаг балок закалочная печь, сек',
       'Скорость прохождения трубы через спрейер, м/с',
       't˚ C трубы после спреера', '1 зона ВТР и уставка отпуск',
       '2 зона ВТР и уставка отпуск', '3 зона ВТР и уставка отпуск',
       '4 зона ВТР и уставка отпуск', '5 зона ВТР и уставка отпуск',
       'шаг балок отпускная печь, сек', 'C', 'Mn', 'Si', 'P', 'S', 'Cr', 'Ni',
       'Cu', 'Al', 'V', 'Ti', 'Nb', 'Mo', 'N', 'B', 'C-coef',
       'Параметр закалка', 'Параметр отпуск', 'Параметр отпуск новый',
       'Параметр отпуск новый 2', 'Параметр отпуск новый V', 'Величина зерна',
       'Тип предела текучести (1186)', 'ICD']

col_out2 = ['Примечание', 'Врем. сопротивление',
       'Предел текучести', '№ партии', 'марка стали', 'диаметр',
       'толщина стенки', 'Гр. прочн.', '1 зона по ВТР закалка',
       '2 зона по ВТР закалка', '3 зона по ВТР закалка',
       'шаг балок закалочная печь, сек',
       'Скорость прохождения трубы через спрейер, м/с',
       't˚ C трубы после спреера', '1 зона ВТР и уставка отпуск',
       '2 зона ВТР и уставка отпуск', '3 зона ВТР и уставка отпуск',
       '4 зона ВТР и уставка отпуск', '5 зона ВТР и уставка отпуск',
       'шаг балок отпускная печь, сек', 'C', 'Mn', 'Si', 'P', 'S', 'Cr', 'Ni',
       'Cu', 'Al', 'V', 'Ti', 'Nb', 'Mo', 'N', 'B', 'C-coef',
       'Параметр закалка', 'Параметр отпуск', 'Параметр отпуск новый',
       'Параметр отпуск новый 2', 'Параметр отпуск новый V', 'Величина зерна',
       'Тип предела текучести (1186)', 'ICD']
#FINISH



load = datetime.now()
database = pd.read_csv('app/DATA/prepared_to_saw_gp.csv', low_memory=False)
print('База исторических режимов загружена за ', datetime.now() - load)

replace_dict_gr = {
    ' ': '',
    '/': '-',
    'ТИП': 'тип',
    'К': 'K',  # русский на английский, везде
    'С': 'C',
    'Р': 'P',
    'Х': 'X',
    'Е': 'E',
    'Т': 'T',
    'М': 'M'
}


def fix_h_group(st):
    st = str(st)
    st = st.upper()
    for it in replace_dict_gr:
        st = st.replace(it, replace_dict_gr[it])
    return st


# возвращает модель, список признаков, по которому строилась модель
def load_model(dir_name):
    model = pickle.load(open(dir_name + '/RF_model_' + '.sav', 'rb'))
    ls_need_col = json.load(open(dir_name + '/ls_need_col', "r"))
    return model, ls_need_col

# ADD START
def close_value(database, col, value):
    database[u'diff'] = np.abs(database[col] - value)
    return database[database[u'diff'] == min(database[u'diff'])][col].values[0]

def max_value(row, val, i):
    tmp = database.copy()
    for col in ls_opt_need:
        zn = {
            'Гр. прочн.': row._5,
            'марка стали': row._2,
            'толщина стенки': row._4,
            'диаметр': row.диаметр,
        }
        z = zn[col]
        tmp2 = tmp[tmp[col] == z]

        if tmp2.shape[0] == 0:
            try:
                tmp = tmp[tmp[col] == close_value(tmp, col, z)]
            except TypeError:
                tmp = pd.DataFrame(columns=col_out2)
                if val == 'u':
                    tmp.loc[0, u'Примечание'] = 'Режим с максимальной скоростью ' + '(' + str(i) + ')'
                else:
                    tmp.loc[0, u'Примечание'] = 'Последний аналогичный режим ' + '(' + str(i) + ')'
                tmp.loc[0, u'№ партии'] = 'Error!!! (причина:' + col + ')'
                return tmp
        else:
            tmp = tmp2.copy()
    if val == 'u':
        tmp[u'summ'] = tmp[u'шаг балок закалочная печь, сек'] + tmp[u'шаг балок отпускная печь, сек']
        df = tmp[tmp[u'summ'] == tmp[u'summ'].min()]
        if df.shape[0] > 10:
            df = df[:10]
        df[u'Примечание'] = None
        df.loc[df[u'summ'] == df[u'summ'].min(), u'Примечание'] = 'Режим с максимальной скоростью ' + '(' + str(i) + ')'
        return df
    else:
        df = tmp[tmp[u'Дата термообработки'] == tmp[u'Дата термообработки'].max()]

        df = df[:1]

        df.loc[df[u'Дата термообработки'] == tmp[
            u'Дата термообработки'].max(), u'Примечание'] = 'Последний аналогичный режим ' + '(' + str(i) + ')'
        return df



def find(database, test):
    database['Гр. прочн.'] = database['Гр. прочн.'].apply(fix_h_group)
    test['Гр. прочн.'] = test['Гр. прочн.'].apply(fix_h_group)
    df1 = pd.DataFrame(columns=database.columns)
    df2 = pd.DataFrame(columns=database.columns)
    for row in test.itertuples():
        df1 = df1.append(max_value(row, 'u', row.Index + 1), sort=False)
        df2 = df2.append(max_value(row, 'p', row.Index + 1), sort=False)

    return df1, df2

def paint_over(first, end, color, sheet_ranges):
    st = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
          'W', 'X', 'Y', 'Z',
          'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR',
          'AS', 'AT', 'AU', 'AV']
    for i in range(first, end + 1):
        for j in st:
            sheet_ranges[j + str(i)].fill = color


def color(direct, ch, out):
    wb = load_workbook(direct)
    sheet_ranges = wb['Sheet1']
    color1 = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    color2 = PatternFill(start_color='00FF7F', end_color='00FF7F', fill_type='solid')
    color3 = PatternFill(start_color='00FF7F', end_color='00FF7F', fill_type='solid')
    k = 0
    for i in range(ch):

        tmp = out[out[u'Примечание'] == 'Предсказание модели ' + '(' + str(i + 1) + ')']
        if tmp.shape[0] == 0:
            continue

        paint_over(2 + k, 2 + k, color1, sheet_ranges)
        tmp2 = out[out[u'Примечание'] == 'Режим с максимальной скоростью ' + '(' + str(i + 1) + ')']

        j = tmp2.shape[0]
        paint_over(3 + k, 2 + j + k, color2, sheet_ranges)
        paint_over(3 + j + k, 3 + j + k, color3, sheet_ranges)
        k = k + j + 2
    wb.save(direct)

def make_result_valid_file(file_name, dir_names, targets, username):
    """Сохраняет файл с результатами """
    result = pd.DataFrame()
    base = pd.DataFrame(columns=database.columns)
    model = []
    ls_need_col = []

    model, ls_need_col = load_model(dir_names)
    time = datetime.now()
    valid = clean_data(file_name, ls_columns_output,username)
    valid.reset_index(drop=True, inplace=True)

    time = datetime.now()

    try:
        y_valid = valid[targets].copy()
    except KeyError:
        y_valid = [0 for x in range(0, valid.shape[0])]
    try:
        X_valid = valid[ls_columns_output].copy()
    except KeyError:
        X_valid = valid[ls_need_col].copy()

    X_valid_c = X_valid[ls_need_col].copy()
    time = datetime.now()

    y_pred = model.predict(X_valid_c)

    for i in range(len(targets)):
        result[targets[i]] = y_valid[targets[i]]
        result['прогноз ' + targets[i]] = y_pred[:, i]
        for x in range(result[targets[i]].shape[0]):
            if result.loc[x, targets[i]] != 0:
                result.loc[x, 'MAE ' + targets[i]] = np.abs(
                    result.loc[x, targets[i]] - result.loc[x, 'прогноз ' + targets[i]])
            else:
                result.loc[x, 'MAE ' + targets[i]] = ''
    time = datetime.now()

    max_u, max_p = find(database, X_valid)
    result = pd.concat([result, X_valid], axis=1)

    return result, max_u, max_p

def main(file,username):
    out = pd.DataFrame(columns=ls_columns_result + ls_columns_output)
    file_name = file
    targets = ['Предел текучести', 'Врем. сопротивление']
    dir_names = os.getcwd() + '/app/DATA/MODELS_RF/H+YS+BATH GS'

    now = datetime.now()
    time = "%d_%ddate %d_%d_%dtime" % (now.day, now.month, now.hour, now.minute, now.second)


    result, max_u, max_p = make_result_valid_file(file_name, dir_names, targets, username)
    for i in range(result.shape[0]):
        result.loc[result.index==i, u'Примечание'] = 'Предсказание модели '+'(' +str(i+1)+')'
    result = result [col_out]
    max_u = calc_all_features(max_u)
    max_p = calc_all_features(max_p)
    max_u = max_u [col_out2]
    max_p = max_p [col_out2]
    for i in range(result.shape[0]):
        out = out.append(result[result.index==i], sort=False)
        out = out.append(max_u[max_u[u'Примечание']=='Режим с максимальной скоростью '+'('+ str(i+1)+')'], sort=False)
        out = out.append(max_p[max_p[u'Примечание']=='Последний аналогичный режим '+'(' +str(i+1)+')'], sort=False)
    out = out[col_out]
    direct = os.getcwd() + '/app/output/' + "prediction_output_" + time + "_" + username + ".xlsx"
    writer = pd.ExcelWriter(direct, engine='xlsxwriter')
    out.to_excel(writer, index=False)
    sh = writer.sheets['Sheet1']
    sh.set_column(0, 0, 42)
    writer.close()
    color(direct, result.shape[0], out)
    writer.close()
#FINISH


